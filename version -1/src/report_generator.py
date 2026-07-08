"""
report_generator.py

Generates Password Security Assessment reports in PDF and Excel formats:
    - Executive Summary Report
    - Password Security Report (per-account detail)
    - Risk Assessment Report
    - Password Policy Compliance Report

Uses reportlab for PDF generation and openpyxl for Excel generation.
"""

from __future__ import annotations

import os
from collections import Counter
from datetime import datetime, timezone
from typing import List

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .assessment_engine import AccountAssessment
from .risk_scorer import RiskLevel


_RISK_COLOR_HEX = {
    RiskLevel.LOW: "2E7D32",
    RiskLevel.MEDIUM: "F9A825",
    RiskLevel.HIGH: "EF6C00",
    RiskLevel.CRITICAL: "C62828",
}


class ReportGenerator:
    """Builds PDF and Excel reports from a list of AccountAssessment results."""

    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        self.styles = getSampleStyleSheet()
        self.styles.add(ParagraphStyle(
            name="ReportTitle", fontSize=20, leading=24, spaceAfter=16, alignment=1,
        ))
        self.styles.add(ParagraphStyle(
            name="SectionHeading", fontSize=14, leading=18, spaceBefore=14, spaceAfter=8,
            textColor=colors.HexColor("#1A237E"),
        ))
        self.styles.add(ParagraphStyle(
            name="SmallNote", fontSize=8, textColor=colors.grey,
        ))

    # ------------------------------------------------------------------
    # Summary statistics helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _summarize(assessments: List[AccountAssessment]) -> dict:
        total = len(assessments)
        risk_counts = Counter(a.risk_result.risk_level.value for a in assessments)
        avg_strength = 0.0
        strength_scores = [a.password_result.strength_score for a in assessments if a.password_result]
        if strength_scores:
            avg_strength = round(sum(strength_scores) / len(strength_scores), 1)
        non_compliant = sum(
            1 for a in assessments if a.policy_result and not a.policy_result.is_compliant
        )
        weak_hashes = sum(
            1 for a in assessments
            if a.hash_result and a.hash_result.security_level.value in ("Deprecated", "Weak")
        )
        return {
            "total": total,
            "risk_counts": risk_counts,
            "avg_strength": avg_strength,
            "non_compliant": non_compliant,
            "weak_hashes": weak_hashes,
        }

    # ------------------------------------------------------------------
    # PDF report
    # ------------------------------------------------------------------
    def generate_pdf_report(self, assessments: List[AccountAssessment], filename: str = "password_security_report.pdf") -> str:
        path = os.path.join(self.output_dir, filename)
        doc = SimpleDocTemplate(
            path, pagesize=letter,
            topMargin=0.6 * inch, bottomMargin=0.6 * inch,
            leftMargin=0.6 * inch, rightMargin=0.6 * inch,
        )
        story = []
        summary = self._summarize(assessments)

        # --- Cover / Executive Summary ---
        story.append(Paragraph("Password Security Assessment Report", self.styles["ReportTitle"]))
        story.append(Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            self.styles["SmallNote"],
        ))
        story.append(Spacer(1, 16))

        story.append(Paragraph("Executive Summary", self.styles["SectionHeading"]))
        exec_lines = [
            f"Total accounts assessed: {summary['total']}",
            f"Average password strength score: {summary['avg_strength']} / 100",
            f"Accounts with policy violations: {summary['non_compliant']}",
            f"Accounts using deprecated/weak hashing algorithms: {summary['weak_hashes']}",
        ]
        for line in exec_lines:
            story.append(Paragraph(line, self.styles["Normal"]))
        story.append(Spacer(1, 10))

        risk_table_data = [["Risk Level", "Account Count"]]
        for level in RiskLevel:
            risk_table_data.append([level.value, str(summary["risk_counts"].get(level.value, 0))])
        risk_table = Table(risk_table_data, colWidths=[3 * inch, 2 * inch])
        risk_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A237E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        story.append(risk_table)
        story.append(PageBreak())

        # --- Risk Assessment Report ---
        story.append(Paragraph("Risk Assessment Report", self.styles["SectionHeading"]))
        risk_detail_data = [["Account", "Risk Level", "Risk Score", "Top Contributing Factor"]]
        for a in assessments:
            top_factor = a.risk_result.contributing_factors[0] if a.risk_result.contributing_factors else "-"
            risk_detail_data.append([
                a.account_identifier,
                a.risk_result.risk_level.value,
                str(a.risk_result.risk_score),
                Paragraph(top_factor, self.styles["Normal"]),
            ])
        risk_detail_table = Table(risk_detail_data, colWidths=[1.3 * inch, 1.2 * inch, 0.8 * inch, 2.7 * inch], repeatRows=1)
        risk_detail_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A237E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ]))
        story.append(risk_detail_table)
        story.append(PageBreak())

        # --- Password Security Report (per account) ---
        story.append(Paragraph("Password Security Report", self.styles["SectionHeading"]))
        for a in assessments:
            story.append(Paragraph(f"Account: {a.account_identifier}", self.styles["Heading3"]))
            if a.password_result:
                pr = a.password_result
                detail = (
                    f"Length: {pr.length} | Entropy: {pr.entropy_bits} bits | "
                    f"Character classes used: {pr.class_count}/4 | "
                    f"Strength: {pr.strength_label} ({pr.strength_score}/100)"
                )
                story.append(Paragraph(detail, self.styles["Normal"]))
                for finding in pr.findings:
                    story.append(Paragraph(f"&bull; {finding}", self.styles["Normal"]))
            if a.hash_result:
                hr = a.hash_result
                story.append(Paragraph(
                    f"Hash Algorithm: {hr.algorithm} | Security Level: {hr.security_level.value} | "
                    f"Salted: {'Yes' if hr.is_salted else 'No'} | Adaptive: {'Yes' if hr.is_adaptive else 'No'}",
                    self.styles["Normal"],
                ))
            if a.policy_result and a.policy_result.violations:
                story.append(Paragraph("Policy Violations:", self.styles["Normal"]))
                for v in a.policy_result.violations:
                    story.append(Paragraph(f"&bull; {v}", self.styles["Normal"]))
            story.append(Spacer(1, 10))
        story.append(PageBreak())

        # --- Recommendations ---
        story.append(Paragraph("Security Recommendations", self.styles["SectionHeading"]))
        for a in assessments:
            if not a.recommendations:
                continue
            story.append(Paragraph(f"Account: {a.account_identifier}", self.styles["Heading3"]))
            rec_data = [["Priority", "Category", "Recommendation"]]
            for r in a.recommendations:
                rec_data.append([r.priority, r.category, Paragraph(r.recommendation, self.styles["Normal"])])
            rec_table = Table(rec_data, colWidths=[0.8 * inch, 1.3 * inch, 3.9 * inch], repeatRows=1)
            rec_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A237E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            story.append(rec_table)
            story.append(Spacer(1, 12))

        doc.build(story)
        return path

    # ------------------------------------------------------------------
    # Excel report
    # ------------------------------------------------------------------
    def generate_excel_report(self, assessments: List[AccountAssessment], filename: str = "password_security_report.xlsx") -> str:
        path = os.path.join(self.output_dir, filename)
        wb = Workbook()

        header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # --- Summary sheet ---
        summary_ws = wb.active
        summary_ws.title = "Executive Summary"
        summary = self._summarize(assessments)
        summary_ws.append(["Password Security Assessment - Executive Summary"])
        summary_ws["A1"].font = Font(size=14, bold=True)
        summary_ws.append([])
        summary_ws.append(["Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")])
        summary_ws.append(["Total Accounts Assessed", summary["total"]])
        summary_ws.append(["Average Password Strength Score", summary["avg_strength"]])
        summary_ws.append(["Accounts with Policy Violations", summary["non_compliant"]])
        summary_ws.append(["Accounts with Weak/Deprecated Hashing", summary["weak_hashes"]])
        summary_ws.append([])
        summary_ws.append(["Risk Level", "Account Count"])
        for cell in summary_ws[summary_ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
        for level in RiskLevel:
            summary_ws.append([level.value, summary["risk_counts"].get(level.value, 0)])
        summary_ws.column_dimensions["A"].width = 34
        summary_ws.column_dimensions["B"].width = 20

        # --- Password Strength sheet ---
        pw_ws = wb.create_sheet("Password Strength")
        pw_headers = [
            "Account", "Length", "Entropy (bits)", "Classes Used", "Common Password?",
            "Sequential Pattern?", "Keyboard Pattern?", "Strength Score", "Strength Label",
        ]
        pw_ws.append(pw_headers)
        for cell in pw_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for a in assessments:
            pr = a.password_result
            if not pr:
                continue
            pw_ws.append([
                a.account_identifier, pr.length, pr.entropy_bits, pr.class_count,
                "Yes" if pr.is_common_password else "No",
                "Yes" if pr.has_sequential_pattern else "No",
                "Yes" if pr.is_keyboard_pattern else "No",
                pr.strength_score, pr.strength_label,
            ])
        self._autofit(pw_ws)

        # --- Hash Analysis sheet ---
        hash_ws = wb.create_sheet("Hash Analysis")
        hash_headers = ["Account", "Algorithm", "Security Level", "Salted?", "Adaptive?", "Cost Factor", "Findings"]
        hash_ws.append(hash_headers)
        for cell in hash_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for a in assessments:
            hr = a.hash_result
            if not hr:
                continue
            hash_ws.append([
                a.account_identifier, hr.algorithm, hr.security_level.value,
                "Yes" if hr.is_salted else "No", "Yes" if hr.is_adaptive else "No",
                hr.cost_factor or "-", "; ".join(hr.findings),
            ])
        self._autofit(hash_ws)

        # --- Policy Compliance sheet ---
        policy_ws = wb.create_sheet("Policy Compliance")
        policy_ws.append(["Account", "Compliant?", "Violations"])
        for cell in policy_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for a in assessments:
            pol = a.policy_result
            if not pol:
                continue
            policy_ws.append([
                a.account_identifier,
                "Yes" if pol.is_compliant else "No",
                "; ".join(pol.violations) if pol.violations else "-",
            ])
        self._autofit(policy_ws)

        # --- Risk Assessment sheet ---
        risk_ws = wb.create_sheet("Risk Assessment")
        risk_ws.append(["Account", "Risk Level", "Risk Score", "Contributing Factors"])
        for cell in risk_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for a in assessments:
            rr = a.risk_result
            risk_ws.append([
                a.account_identifier, rr.risk_level.value, rr.risk_score,
                "; ".join(rr.contributing_factors),
            ])
        self._autofit(risk_ws)

        # --- Recommendations sheet ---
        rec_ws = wb.create_sheet("Recommendations")
        rec_ws.append(["Account", "Priority", "Category", "Recommendation"])
        for cell in rec_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for a in assessments:
            for r in a.recommendations:
                rec_ws.append([a.account_identifier, r.priority, r.category, r.recommendation])
        self._autofit(rec_ws)

        wb.save(path)
        return path

    @staticmethod
    def _autofit(ws, max_width: int = 60):
        for col_cells in ws.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(length + 2, 12), max_width)
