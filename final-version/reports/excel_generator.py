"""
ExcelReportGenerator
====================
Generates Excel (.xlsx) reports using openpyxl.
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


RISK_FILL = {
    "Low": "C6EFCE",
    "Medium": "FFEB9C",
    "High": "FFD1A3",
    "Critical": "FFC7CE",
}


class ExcelReportGenerator:

    def __init__(self, output_dir="generated_reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_assessment_report(self, assessments: list, report_title="Password_Security_Report") -> str:
        filename = f"{report_title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Assessments"

        headers = ["Username", "Department", "Strength Score", "Entropy (bits)", "Risk Level",
                   "Risk Score", "Hash Algorithm", "Deprecated Hash", "Policy Compliant",
                   "Common Password", "Reused", "Assessed At"]
        ws.append(headers)

        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for a in assessments:
            row = [
                a.get("username", ""),
                a.get("department", "") or "-",
                a.get("strength_score", ""),
                float(a.get("entropy_bits") or 0),
                a.get("risk_level", ""),
                a.get("risk_score", ""),
                a.get("hash_algorithm", ""),
                "Yes" if a.get("hash_deprecated") else "No",
                "Yes" if a.get("policy_compliant") else "No",
                "Yes" if a.get("is_common_password") else "No",
                "Yes" if a.get("is_reused") else "No",
                str(a.get("assessed_at", "")),
            ]
            ws.append(row)
            risk_row = ws.max_row
            fill_color = RISK_FILL.get(a.get("risk_level"))
            if fill_color:
                ws.cell(row=risk_row, column=5).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 2)

        wb.save(filepath)
        return filepath

    def generate_policy_compliance_report(self, assessments: list, violations_map: dict) -> str:
        filename = f"Policy_Compliance_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Compliance"
        ws.append(["Username", "Compliant", "Violations"])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for a in assessments:
            violations = violations_map.get(a.get("assessment_id"), [])
            ws.append([
                a.get("username", ""),
                "Yes" if a.get("policy_compliant") else "No",
                "; ".join(v["type"] for v in violations) if violations else "None",
            ])

        for i, width in enumerate([20, 12, 60], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(filepath)
        return filepath
